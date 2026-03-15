from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from typing import Any

import pandas as pd
import torch
from difflib import get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.workbook.workbook import Workbook
from transformers import AutoModelForCausalLM, AutoTokenizer

# =========================================================
# CONFIG
# =========================================================
DEBUG = True
MODEL_NAME = "Qwen/Qwen2.5-0.5B-Instruct"

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

SYSTEM_PROMPT = """
Você é um assistente de planilhas.
Sua função é interpretar pedidos do usuário e retornar SOMENTE um JSON válido.
Não escreva explicações fora do JSON.
Nunca use markdown.
Nunca use blocos de código.

Ações permitidas:
- destacar_coluna
- explicar_coluna
- buscar_texto
- resumir_aba
- listar_abas
- listar_colunas

Formato esperado:
{
  "acao": "nome_da_acao",
  "aba": "nome_da_aba_ou_null",
  "coluna": "nome_da_coluna_ou_null",
  "texto": "texto_ou_null"
}

Regras:
- Se o usuário pedir para destacar uma coluna, use "destacar_coluna".
- Se o usuário pedir explicação sobre uma coluna, use "explicar_coluna".
- Se o usuário pedir para procurar um protocolo, código, texto ou termo, use "buscar_texto".
- Se o usuário pedir um resumo da aba, use "resumir_aba".
- Se o usuário pedir abas disponíveis, use "listar_abas".
- Se o usuário pedir colunas disponíveis, use "listar_colunas".
- Se não souber a aba, retorne aba como null.
- Se não souber a coluna, retorne coluna como null.
- Se não houver texto, retorne texto como null.
- Responda sempre em JSON válido.
""".strip()


def lista_arquivos() -> dict[int, str]:
    return {
        i: str(arquivo)
        for i, arquivo in enumerate(
            [
                arq for arq in INPUT_DIR.iterdir()
                if arq.is_file() and arq.suffix.lower() in {".csv", ".xlsx"}
            ],
            start=1
        )
    }


def ler_arquivo(caminho_arquivo: str) -> tuple[pd.DataFrame | dict[str, pd.DataFrame], Workbook | None]:
    if caminho_arquivo.endswith(".csv"):
        return pd.read_csv(caminho_arquivo), None
    if caminho_arquivo.endswith(".xlsx"):
        return pd.read_excel(caminho_arquivo, sheet_name=None), load_workbook(caminho_arquivo)
    raise ValueError("Formato de arquivo não suportado")


def normalizar_planilhas(dados: pd.DataFrame | dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    if isinstance(dados, pd.DataFrame):
        return {"Planilha1": dados}
    return dados


def encontrar_nome_proximo(opcoes: list[str], termo: str | None, cutoff: float = 0.5) -> str | None:
    if termo is None:
        return None

    opcoes_str = [str(x) for x in opcoes]
    termo_lower = termo.lower().strip()

    for opc in opcoes_str:
        if opc.lower().strip() == termo_lower:
            return opc

    for opc in opcoes_str:
        if termo_lower in opc.lower():
            return opc

    matches = get_close_matches(termo_lower, [x.lower() for x in opcoes_str], n=1, cutoff=cutoff)
    if not matches:
        return None

    for opc in opcoes_str:
        if opc.lower() == matches[0]:
            return opc
    return None


def extrair_json_bruto(texto: str) -> dict[str, Any]:
    texto = texto.strip()

    try:
        return json.loads(texto)
    except Exception:
        pass

    match = re.search(r"\{.*\}", texto, re.DOTALL)
    if match:
        trecho = match.group(0)
        try:
            return json.loads(trecho)
        except Exception:
            pass

    raise ValueError(f"Não foi possível extrair JSON válido da resposta do modelo:\n{texto}")


def detectar_device() -> str:
    return "cuda" if torch.cuda.is_available() else "cpu"


class SpreadsheetLLMAgent:
    def __init__(self, model_name: str = MODEL_NAME):
        self.device = detectar_device()

        if DEBUG:
            print(f"[DEBUG] Carregando modelo: {model_name}")
            print(f"[DEBUG] Device: {self.device}")

        self.tokenizer = AutoTokenizer.from_pretrained(model_name, trust_remote_code=True)

        if self.device == "cuda":
            self.model = AutoModelForCausalLM.from_pretrained(
                model_name,
                dtype="auto",
                device_map="auto",
                trust_remote_code=True
            )
        else:
            self.model = AutoModelForCausalLM.from_pretrained(
                model_name,
                dtype=torch.float32,
                trust_remote_code=True
            )
            self.model.to(self.device)

    def interpretar(
        self,
        comando_usuario: str,
        nomes_abas: list[str],
        colunas_por_aba: dict[str, list[str]]
    ) -> dict[str, Any]:
        contexto = {
            "abas_disponiveis": nomes_abas,
            "colunas_por_aba": colunas_por_aba,
            "pedido_usuario": comando_usuario
        }

        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": json.dumps(contexto, ensure_ascii=False, indent=2)}
        ]

        prompt = self.tokenizer.apply_chat_template(
            messages,
            tokenize=False,
            add_generation_prompt=True
        )

        inputs = self.tokenizer(prompt, return_tensors="pt").to(self.model.device)
        outputs = self.model.generate(
            **inputs,
            max_new_tokens=220,
            do_sample=False,
            temperature=0.0
        )

        generated_ids = outputs[0][inputs["input_ids"].shape[1]:]
        resposta = self.tokenizer.decode(generated_ids, skip_special_tokens=True).strip()

        if DEBUG:
            print("\n[DEBUG] Resposta bruta do modelo:")
            print(resposta)

        acao = extrair_json_bruto(resposta)
        acao.setdefault("acao", None)
        acao.setdefault("aba", None)
        acao.setdefault("coluna", None)
        acao.setdefault("texto", None)

        return acao


def criar_copia_para_output(caminho_arquivo: str, nome_base_saida: str) -> Path:
    origem = Path(caminho_arquivo)
    destino = OUTPUT_DIR / f"{nome_base_saida}.xlsx"

    if origem.suffix.lower() == ".xlsx":
        shutil.copy2(origem, destino)
    elif origem.suffix.lower() == ".csv":
        df = pd.read_csv(origem)
        with pd.ExcelWriter(destino, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Planilha1", index=False)
    else:
        raise ValueError("Formato não suportado para cópia.")

    return destino


def destacar_coluna_excel(wb: Workbook, nome_aba: str, nome_coluna: str):
    ws = wb[nome_aba]
    cabecalhos = [cell.value for cell in ws[1]]

    if nome_coluna not in cabecalhos:
        raise ValueError(f"Coluna '{nome_coluna}' não encontrada na aba '{nome_aba}'.")

    idx_coluna = cabecalhos.index(nome_coluna) + 1

    fill = PatternFill(fill_type="solid", fgColor="FFF59D")
    header_fill = PatternFill(fill_type="solid", fgColor="FFD54F")
    bold_font = Font(bold=True)

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=idx_coluna)
        cell.fill = fill
        if row == 1:
            cell.fill = header_fill
            cell.font = bold_font


def explicar_coluna(df: pd.DataFrame, nome_coluna: str) -> str:
    if nome_coluna not in df.columns:
        return f"A coluna '{nome_coluna}' não foi encontrada."

    serie = df[nome_coluna]
    texto = []
    texto.append(f"Coluna analisada: {nome_coluna}")
    texto.append(f"Quantidade de registros: {len(serie)}")
    texto.append(f"Valores nulos: {int(serie.isna().sum())}")

    if pd.api.types.is_numeric_dtype(serie):
        texto.append("Tipo: numérico")
        texto.append(f"Menor valor: {serie.min()}")
        texto.append(f"Maior valor: {serie.max()}")
        texto.append(f"Média: {serie.mean():.2f}")
        texto.append(f"Soma: {serie.sum():.2f}")
    else:
        texto.append("Tipo: texto/categórico")
        exemplos = list(serie.dropna().astype(str).unique()[:10])
        texto.append(f"Exemplos de valores: {exemplos}")

    return "\n".join(texto)


def buscar_texto_em_dataframe(df: pd.DataFrame, texto: str) -> pd.DataFrame:
    mask = df.astype(str).apply(lambda col: col.str.contains(texto, case=False, na=False))
    return df[mask.any(axis=1)]


def resumir_aba(df: pd.DataFrame, nome_aba: str) -> str:
    linhas, colunas = df.shape
    texto = [f"Resumo da aba '{nome_aba}':"]
    texto.append(f"- Linhas: {linhas}")
    texto.append(f"- Colunas: {colunas}")
    texto.append(f"- Nomes das colunas: {list(df.columns)}")

    numericas = df.select_dtypes(include="number").columns.tolist()
    categoricas = [c for c in df.columns if c not in numericas]

    texto.append(f"- Colunas numéricas: {numericas}")
    texto.append(f"- Colunas categóricas/textuais: {categoricas}")

    if numericas:
        texto.append("- Estatísticas básicas das colunas numéricas:")
        for col in numericas[:10]:
            texto.append(f"  * {col}: min={df[col].min()}, max={df[col].max()}, média={df[col].mean():.2f}")

    return "\n".join(texto)


def resolver_aba(planilhas: dict[str, pd.DataFrame], aba_sugerida: str | None) -> str:
    abas = list(planilhas.keys())

    if len(abas) == 1:
        return abas[0]

    encontrada = encontrar_nome_proximo(abas, aba_sugerida)
    if encontrada:
        return encontrada

    return abas[0]


def resolver_coluna(df: pd.DataFrame, coluna_sugerida: str | None, fallback_keywords: list[str] | None = None) -> str | None:
    colunas = [str(c) for c in df.columns]

    encontrada = encontrar_nome_proximo(colunas, coluna_sugerida)
    if encontrada:
        return encontrada

    if fallback_keywords:
        for kw in fallback_keywords:
            encontrada = encontrar_nome_proximo(colunas, kw)
            if encontrada:
                return encontrada

    return None


def executar_acao(
    acao: dict[str, Any],
    caminho_arquivo: str,
    planilhas: dict[str, pd.DataFrame],
    wb: Workbook | None
):
    nome_arquivo_base = Path(caminho_arquivo).stem
    tipo_acao = acao.get("acao")
    aba_sugerida = acao.get("aba")
    coluna_sugerida = acao.get("coluna")
    texto_sugerido = acao.get("texto")

    if tipo_acao == "listar_abas":
        print("Abas disponíveis:")
        for aba in planilhas.keys():
            print(f"- {aba}")
        return

    if tipo_acao == "listar_colunas":
        aba_real = resolver_aba(planilhas, aba_sugerida)
        print(f"Colunas da aba '{aba_real}':")
        for col in planilhas[aba_real].columns:
            print(f"- {col}")
        return

    if tipo_acao == "resumir_aba":
        aba_real = resolver_aba(planilhas, aba_sugerida)
        print(resumir_aba(planilhas[aba_real], aba_real))
        return

    if tipo_acao == "explicar_coluna":
        aba_real = resolver_aba(planilhas, aba_sugerida)
        df = planilhas[aba_real]

        coluna_real = resolver_coluna(
            df,
            coluna_sugerida,
            fallback_keywords=["lucro", "profit", "margem", "resultado", "faturamento", "receita"]
        )

        if coluna_real is None:
            print("Não foi possível identificar a coluna para explicar.")
            return

        print(explicar_coluna(df, coluna_real))
        return

    if tipo_acao == "buscar_texto":
        aba_real = resolver_aba(planilhas, aba_sugerida)
        df = planilhas[aba_real]

        if not texto_sugerido:
            print("Nenhum texto foi informado para busca.")
            return

        encontrados = buscar_texto_em_dataframe(df, texto_sugerido)
        print(f"Foram encontradas {len(encontrados)} linha(s) com o texto '{texto_sugerido}' na aba '{aba_real}'.")

        if len(encontrados) > 0:
            print(encontrados.head(20).to_string(index=False))
            caminho_saida = OUTPUT_DIR / f"{nome_arquivo_base}_busca_{re.sub(r'[^a-zA-Z0-9_-]', '_', texto_sugerido)}.xlsx"
            encontrados.to_excel(caminho_saida, index=False)
            print(f"\nResultado salvo em: {caminho_saida}")

        return

    if tipo_acao == "destacar_coluna":
        aba_real = resolver_aba(planilhas, aba_sugerida)
        df = planilhas[aba_real]

        coluna_real = resolver_coluna(
            df,
            coluna_sugerida,
            fallback_keywords=["lucro", "profit", "margem", "resultado", "faturamento", "receita"]
        )

        if coluna_real is None:
            print("Não foi possível identificar a coluna a ser destacada.")
            return

        nome_saida = f"{nome_arquivo_base}_destacado_{re.sub(r'[^a-zA-Z0-9_-]', '_', coluna_real)}"
        caminho_saida = criar_copia_para_output(caminho_arquivo, nome_saida)

        wb_saida = load_workbook(caminho_saida)
        aba_wb = encontrar_nome_proximo(wb_saida.sheetnames, aba_real) or wb_saida.sheetnames[0]

        destacar_coluna_excel(wb_saida, aba_wb, coluna_real)
        wb_saida.save(caminho_saida)

        print(f"Coluna '{coluna_real}' destacada com sucesso na aba '{aba_wb}'.")
        print(f"Arquivo gerado em: {caminho_saida}")
        return

    print("Ação não reconhecida ou não suportada.")
    print("JSON retornado pelo modelo:")
    print(json.dumps(acao, ensure_ascii=False, indent=2))


def main():
    arquivos = lista_arquivos()

    if not arquivos:
        print("Nenhum arquivo encontrado na pasta 'inputs'.")
        return

    print("Arquivos disponíveis:")
    for idx, caminho in arquivos.items():
        print(f"{idx}: {caminho}")

    try:
        escolha = int(input("\nDigite o número do arquivo que deseja ler: ").strip())
    except ValueError:
        print("Entrada inválida.")
        return

    caminho_arquivo = arquivos.get(escolha)
    if caminho_arquivo is None:
        print("Arquivo inválido.")
        return

    dados, wb = ler_arquivo(caminho_arquivo)
    planilhas = normalizar_planilhas(dados)

    print("\nArquivo carregado com sucesso.")
    print("Abas disponíveis:")
    for nome_aba, df in planilhas.items():
        print(f"- {nome_aba} ({df.shape[0]} linhas, {df.shape[1]} colunas)")

    colunas_por_aba = {
        aba: [str(c) for c in df.columns]
        for aba, df in planilhas.items()
    }

    agente = SpreadsheetLLMAgent()

    while True:
        comando = input("\nDigite seu comando (ou 'sair'): ").strip()
        if comando.lower() in {"sair", "exit", "quit"}:
            print("Encerrando assistente.")
            break

        try:
            acao = agente.interpretar(
                comando_usuario=comando,
                nomes_abas=list(planilhas.keys()),
                colunas_por_aba=colunas_por_aba
            )

            if DEBUG:
                print("\n[DEBUG] JSON interpretado:")
                print(json.dumps(acao, ensure_ascii=False, indent=2))

            executar_acao(acao, caminho_arquivo, planilhas, wb)

        except Exception as e:
            print(f"Erro ao processar comando: {e}")


if __name__ == "__main__":
    main()